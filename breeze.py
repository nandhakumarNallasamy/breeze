from breeze_connect import BreezeConnect
from datetime import datetime, timedelta, timezone
import logging
import pytz
from IPython.display import clear_output
import os
import threading
from time import sleep
from concurrent.futures import ThreadPoolExecutor
import concurrent.futures
import pickle
import pandas as pd

# Global Variables
breeze = None
logger = None
contract_registry = {}
token_to_contract_map = {}
ORDER_THREAD_POOL = ThreadPoolExecutor(max_workers=10)


# Pickle files
PICKLE_CONNECTION_FILE = 'breeze_connection.pkl'
PICKLE_TOKEN_MAP_FILE = 'token_map.pkl'

def cleanup_pkl_files(directory='.', max_age_hours=8):
    # Delete pickle files older than specified hours
    current_time = datetime.now()
    
    for filename in os.listdir(directory):
        if filename.endswith('.pkl'):
            file_path = os.path.join(directory, filename)
            file_modification_time = datetime.fromtimestamp(os.path.getmtime(file_path))
            
            # Calculate file age
            file_age = current_time - file_modification_time
            
            # Delete if file is older than max_age_hours
            if file_age > timedelta(hours=max_age_hours):
                try:
                    os.remove(file_path)
                    print(f"Deleted {filename} (older than {max_age_hours} hours)")
                except Exception as e:
                    print(f"Error deleting {filename}: {e}")
                    
def connect(session_token=None):
    # Initialize and connect to BreezeConnect API with WebSocket
    global breeze
    
    # Try to load existing connection from pickle
    try:
        if os.path.exists(PICKLE_CONNECTION_FILE):
            with open(PICKLE_CONNECTION_FILE, 'rb') as f:
                breeze = pickle.load(f)
            logger.info("✅ Loaded existing Breeze connection from pickle")
        
        else:
            raise Exception("No connection pickle found")
                
    except Exception as e:
        logger.warning(f"❌ Failed to load pickled connection: {e}")
        
        # If no pickled connection, create new
        api_key = os.environ.get("BREEZE_API_KEY")
        api_secret = os.environ.get("BREEZE_API_SECRET_KEY")
        breeze = BreezeConnect(api_key)
        
        if session_token is None:
            from session_token import get_icici_session_token
            session_token = get_icici_session_token()
        
        try:
            breeze.generate_session(api_secret, session_token)
            logger.info("✅ Successfully connected to BreezeConnect")
            
            try:
                with open(PICKLE_CONNECTION_FILE, 'wb') as f:
                    pickle.dump(breeze, f)
                    logger.info("✅ Saved Breeze connection to pickle")
            except Exception as e:
                logger.error(f"❌ Failed to save connection to pickle: {e}")
        
        except Exception as e:
            logger.error(f"❌ Failed to generate session: {e}")
            exit(1)
    
    breeze.ws_disconnect()    
    def on_ticks(ticks):
        # Process incoming websocket data for both tick and OHLCV data
        try:
            # CASE 1: OHLCV data (identified by 'interval' field)
            if 'interval' in ticks:
                # Extract common data
                stock_code = ticks.get('stock_code')
                
                # Create OHLCV candle data structure
                candle_data = {
                    'datetime': ticks.get('datetime'),
                    'open': float(ticks.get('open')),
                    'high': float(ticks.get('high')),
                    'low': float(ticks.get('low')),
                    'close': float(ticks.get('close')),
                    'volume': int(ticks.get('volume', 0))
                }
                
                # Add open interest if available
                if 'oi' in ticks:
                    candle_data['open_interest'] = int(ticks.get('oi', 0))
                
                expiry_date = None    
                if 'expiry_date' in ticks:
                    expiry_date = ticks.get('expiry_date')  # Format: "DD-MMM-YYYY"
                
                # Handle options data (has right_type field)
                if 'right_type' in ticks and ticks.get('strike_price'):
                    # Get options-specific data
                    strike_price = int(float(ticks.get('strike_price')))
                    right_type = ticks.get('right_type')  # CE or PE
                    
                    # Convert right_type to expected format
                    right = 'call' if right_type == 'CE' else 'put'
                    
                    # Find matching options contract using flexible approach
                    contract = None
                    
                    # Loop through registry to find the right contract
                    for reg_key, reg_contract in contract_registry.items():
                        # Check if this is the correct contract by matching:
                        # 1. The stock code (e.g., "NIFTY")
                        # 2. The strike price (e.g., "22000")
                        # 3. The right type (call/put)
                        if (stock_code in reg_key and str(strike_price).split('.')[0] in reg_key 
                            and right in reg_key.lower() and expiry_date in reg_key):
                            contract = reg_contract
                            break
                    
                    # If found, update the contract
                    if contract is not None and contract.is_ohlcv_subscribed:
                        # Add the candle to the contract's data
                        contract.ohlcv_data.append(candle_data)
                    else:
                        logger.error(f"❌ Error finding matching contract for {stock_code}-{expiry_date}-{strike_price}-{right}(OHLCV)")
                
                # Handle futures data
                else:
                    # Find matching futures contract
                    contract = None
                    
                    # Loop through registry to find futures contract
                    for reg_key, reg_contract in contract_registry.items():
                        if stock_code in reg_key and "FUT" in reg_key and expiry_date in reg_key:
                            contract = reg_contract
                            break
                    
                    # If found, update the contract
                    if contract:
                        # Add the candle to the contract's data
                        contract.ohlcv_data.append(candle_data)
                    else:
                        logger.error(f"❌ Error finding matching contract for {stock_code}-FUT-{expiry_date}(OHLCV)")
            
            # CASE 2: Tick data (has 'last' field but no 'interval')
            elif 'last' in ticks:
                # Add this part to handle symbol-based tick data
                if 'symbol' in ticks:
                    # Extract the token ID from the symbol (e.g., from '4.1!35559' get '35559')
                    token_id = ticks['symbol'].split('!')[1] if '!' in ticks['symbol'] else None
                    
                    if token_id and token_id in token_to_contract_map:
                        # Get contract info from the mapping
                        contract_info = token_to_contract_map[token_id]
                        
                        # Create registry key based on the contract details already parsed
                        registry_key = None
                        
                        if contract_info.get('product_type') == 'options':
                            # Format: "{stock_code}-{expiry_date}-{strike_price}-{right}"
                            registry_key = f"{contract_info['stock_code']}-{contract_info['expiry_date']}-{contract_info['strike_price']}-{contract_info['right']}"
                        elif contract_info.get('product_type') == 'futures':
                            # Format: "{stock_code}-FUT-{expiry_date}"
                            registry_key = f"{contract_info['stock_code']}-FUT-{contract_info['expiry_date']}"
                        
                        # Look up contract in registry
                        contract = None
                        if registry_key in contract_registry:
                            contract = contract_registry[registry_key]
                        else:
                            # Try a more flexible lookup
                            for reg_key, reg_contract in contract_registry.items():
                                if contract_info['stock_code'] in reg_key:
                                    if contract_info.get('product_type') == 'options':
                                        if (contract_info['strike_price'] in reg_key and 
                                            contract_info['right'] in reg_key.lower() and 
                                            contract_info['expiry_date'] in reg_key):
                                            contract = reg_contract
                                            break
                                    elif contract_info.get('product_type') == 'futures':
                                        if "FUT" in reg_key and contract_info['expiry_date'] in reg_key:
                                            contract = reg_contract
                                            break
                        
                        # Update contract data if found
                        if contract is not None and contract.is_ltp_subscribed:
                            contract.ltp = float(ticks['last'])
                            contract.last_update_time = datetime.now()
                        else:
                            logger.debug(f"❌ Error finding matching contract for {stock_code}-{expiry_date}-{strike_price}-{right}(Tick-by_Tick)")
            
            # CASE 3: Market Depth data (identified by 'depth')
            elif 'depth' in ticks:
                # Extract the token from the symbol
                token_id = ticks['symbol'].split('!')[1] if '!' in ticks['symbol'] else None
                
                if token_id and token_id in token_to_contract_map:
                        # Get contract info from the mapping
                        contract_info = token_to_contract_map[token_id]
                        
                        # Create registry key based on the contract details already parsed
                        registry_key = None
                        
                        if contract_info.get('product_type') == 'options':
                            # Format: "{stock_code}-{expiry_date}-{strike_price}-{right}"
                            registry_key = f"{contract_info['stock_code']}-{contract_info['expiry_date']}-{contract_info['strike_price']}-{contract_info['right']}"
                        elif contract_info.get('product_type') == 'futures':
                            # Format: "{stock_code}-FUT-{expiry_date}"
                            registry_key = f"{contract_info['stock_code']}-FUT-{contract_info['expiry_date']}"
                            
                # Look up contract in registry
                        contract = None
                        if registry_key in contract_registry:
                            contract = contract_registry[registry_key]
                        else:
                            # Try a more flexible lookup
                            for reg_key, reg_contract in contract_registry.items():
                                if contract_info['stock_code'] in reg_key:
                                    if contract_info.get('product_type') == 'options':
                                        if (contract_info['strike_price'] in reg_key and 
                                            contract_info['right'] in reg_key.lower() and 
                                            contract_info['expiry_date'] in reg_key):
                                            contract = reg_contract
                                            break
                                    elif contract_info.get('product_type') == 'futures':
                                        if "FUT" in reg_key and contract_info['expiry_date'] in reg_key:
                                            contract = reg_contract
                                            break
                
                # If found, update market depth data
                if contract is not None and contract.is_depth_subscribed:
                    # Process depth data
                    if ticks['depth'] and len(ticks['depth']) > 0:
                        # Update best bid/ask information
                        first_level = ticks['depth'][0]
                        
                        contract.best_bid_price = float(first_level.get('BestBuyRate-1', 0))
                        contract.best_bid_qty = int(first_level.get('BestBuyQty-1', 0))
                        contract.best_ask_price = float(first_level.get('BestSellRate-1', 0))
                        contract.best_ask_qty = int(first_level.get('BestSellQty-1', 0))
                        contract.depth_update_time = datetime.now()
                        
                        # Optional: store all depth levels
                        contract.bids = []
                        contract.asks = []
                        
                        for level in ticks['depth']:
                            for i in range(1, 6):  # Process all 5 levels
                                # Get bid data
                                bid_price_key = f'BestBuyRate-{i}'
                                bid_qty_key = f'BestBuyQty-{i}'
                                bid_orders_key = f'BuyNoOfOrders-{i}'
                                
                                # Process bid if data exists
                                if bid_price_key in level and bid_qty_key in level:
                                    bid_price = float(level.get(bid_price_key, 0))
                                    bid_qty = int(level.get(bid_qty_key, 0))
                                    bid_orders = int(level.get(bid_orders_key, 0))
                                    contract.bids.append([bid_price, bid_qty, bid_orders])
                                
                                # Get ask data
                                ask_price_key = f'BestSellRate-{i}'
                                ask_qty_key = f'BestSellQty-{i}'
                                ask_orders_key = f'SellNoOfOrders-{i}'
                                
                                # Process ask if data exists
                                if ask_price_key in level and ask_qty_key in level:
                                    ask_price = float(level.get(ask_price_key, 0))
                                    ask_qty = int(level.get(ask_qty_key, 0))
                                    ask_orders = int(level.get(ask_orders_key, 0))
                                    contract.asks.append([ask_price, ask_qty, ask_orders])
            
        except Exception as e:
            logger.error(f"❌ Error in websocket callback: {e}")
            import traceback
            traceback.print_exc()
    
    breeze.on_ticks = on_ticks
    breeze.ws_connect()
    logger.info("✅ Websocket connected")
    
    return breeze

def get_breeze():
    # Check if .pkl files are recent
    cleanup_pkl_files()
    
    # Load the token mapping csv
    load_token_mapping()
    
    # Connect to Breeze API and websocket
    connect()
 
def get_ist_time():
    # Get IST
    utc_now = datetime.now(timezone.utc)
    ist_now = utc_now + timedelta(hours=5, minutes=30)
    return ist_now.replace(tzinfo=None)

def get_iso_date(dateString=None):
    # Get ISO date
    if dateString is None:
        return get_ist_time().strftime("%Y-%m-%d") + 'T06:00:00.000Z'
    
    try:
        date_obj = datetime.strptime(dateString, "%d-%b-%Y")
        return date_obj.strftime("%Y-%m-%d") + 'T06:00:00.000Z'
    except ValueError:
        logger.warning(f"❗ Invalid date format: {dateString}, using current date")
        return get_ist_time().strftime("%Y-%m-%d") + 'T06:00:00.000Z'

def get_iso_datetime(dateString=None):
    # Get ISO dateTime
    if dateString is None:
        return get_ist_time().strftime("%Y-%m-%dT%H:%M:%S.000Z")
    
    try:
        date_obj = datetime.strptime(dateString, "%d-%b-%Y %H:%M:%S")
        return date_obj.strftime("%Y-%m-%dT%H:%M:%S.000Z")
    except ValueError:
        logger.warning(f"❗ Invalid datetime format: {dateString}, using current time")
        return get_ist_time().strftime("%Y-%m-%dT%H:%M:%S.000Z")

def convert_iso_to_breeze_date(iso_date):
    # Convert ISO format date to Breeze API format (DD-MMM-YYYY)
    # Input: "2025-03-27T06:00:00.000Z" or "2025-03-27"
    # Output: "27-Mar-2025"
    try:
        if 'T' in iso_date:
            iso_date = iso_date.split('T')[0]
        date_obj = datetime.strptime(iso_date, "%Y-%m-%d")
        return date_obj.strftime("%d-%b-%Y")
    except ValueError:
        if logger:
            logger.warning(f"❗ Invalid ISO date format: {iso_date}")
        return None

def log():
    # Set up logging
    logging.basicConfig(level = logging.INFO, format = '%(asctime)s - %(levelname)s - %(message)s')
    logger = logging.getLogger(__name__)
    return logger

# Function to download and prepare the token to contract details mapping
def load_token_mapping(url="https://traderweb.icicidirect.com/Content/File/txtFile/ScripFile/StockScriptNew.csv"):
    import requests
    from io import StringIO
    global token_to_contract_map
    
    try:
        if os.path.exists(PICKLE_TOKEN_MAP_FILE):
            with open(PICKLE_TOKEN_MAP_FILE, 'rb') as f:
                token_to_contract_map = pickle.load(f)
            logger.info(f"✅ Loaded {len(token_to_contract_map)} token mappings from pickle")
            return True
        else:
            raise Exception("No token_to_contract_map pickle found")
    except Exception as e:
        logger.error(f"❌ Error loading token map from pickle: {e}")
        
        try:
            logger.info(f"🌐 Downloading stock script CSV from {url}")
            response = requests.get(url)
            df = pd.read_csv(StringIO(response.text))
            
            for _, row in df.iterrows():
                # Only process rows with valid TK values
                if pd.notna(row['TK']) and row['TK'] != 0:
                    # Convert TK to string to ensure consistent lookup
                    tk = str(row['TK'])
                    
                    
                    # Get stock code from SM field
                    sm = str(row['SM'])
                    
                    # Parse SM field to extract all contract details
                    if '~' in sm:
                        parts = sm.split('~')
                        stock_code = parts[0]
                        
                        if len(parts) > 1 and ':' in parts[1]:
                            # Parse contract specification (e.g., "O:18-Mar-2025:CE:7400000")
                            contract_specs = parts[1].split(':')
                            
                            if len(contract_specs) >= 2:
                                contract_type = contract_specs[0]  # 'O' for options, 'F' for futures
                                expiry_date = contract_specs[1]    # Format: DD-MMM-YYYY
                                
                                if contract_type == 'O' and len(contract_specs) >= 4:
                                    # This is an options contract
                                    right_code = contract_specs[2]  # 'CE' or 'PE'
                                    right = 'call' if right_code == 'CE' else 'put'
                                    3
                                    # Strike price needs to be divided by 100
                                    strike_raw = contract_specs[3]
                                    try:
                                        strike_price = str(int(int(strike_raw) / 100))
                                    except (ValueError, TypeError):
                                        strike_price = strike_raw
                                    
                                    token_to_contract_map[tk] = {
                                        'stock_code': stock_code,
                                        'expiry_date': expiry_date,
                                        'strike_price': strike_price,
                                        'right': right,
                                        'product_type': 'options'
                                    }
                                    
                                elif contract_type == 'F':
                                    # This is a futures 2.6ontract
                                    token_to_contract_map[tk] = {
                                        'stock_code': stock_code,
                                        'expiry_date': expiry_date,
                                        'product_type': 'futures'
                                    }
                    else:
                        # If SM doesn't have extended info, just store the stock code
                        token_to_contract_map[tk] = {'stock_code': sm}
            
            logger.info(f"✅ Loaded {len(token_to_contract_map)} token mappings")
            
            try:
                with open(PICKLE_TOKEN_MAP_FILE, 'wb') as f:
                    pickle.dump(token_to_contract_map, f)
                    logger.info(f"✅ Tokken mappings pickled")
            except Exception as e:
                logger.error(f"❌ Failed to pickle token mappings: {e}")
                
            return True
        except Exception as e:
            logger.error(f"❌ Error loading token mappings: {e}")
            return False
        
def contract_exists(my_contract):
    #Check if a contract exists in the Breeze dictionary mapping
    
    # Get dictionary index based on exchange code
    exchange_to_index = {
        "BSE": 0, "NSE": 1, "NDX": 2, "MCX": 3, "NFO": 4, "BFO": 5
    }
    target_index = exchange_to_index.get(my_contract.exchange_code.upper(), None)
    
    if target_index is None or not hasattr(breeze, 'stock_script_dict_list'):
        return False, None
    
    # For F&O contracts
    if hasattr(my_contract, 'product_type') and my_contract.product_type:
        if my_contract.product_type.lower() == "options":
            # Use the existing convert_iso_to_breeze_date function
            expiry_date = convert_iso_to_breeze_date(my_contract.expiry_date) if hasattr(my_contract, 'expiry_date') else ""
            right_code = "CE" if my_contract.right.lower() == "call" else "PE"
            contract_key = f"OPT-{my_contract.stock_code}-{expiry_date}-{my_contract.strike_price}-{right_code}"
        else:  # futures
            expiry_date = convert_iso_to_breeze_date(my_contract.expiry_date) if hasattr(my_contract, 'expiry_date') else ""
            contract_key = f"FUT-{my_contract.stock_code}-{expiry_date}"
    else:  # equity
        contract_key = my_contract.stock_code
    
    # Check if contract exists
    exists = contract_key in breeze.stock_script_dict_list[target_index]
    return exists

def subscribe_feed(contract, interval=None, get_exchange_quotes=True, get_market_depth=False):
    #Subscribe to real-time feed for a contract
    
    exists = contract_exists(contract)
    if not exists:
        print(f"⚠️ Contract not found in dictionary: {contract.short_hand}")
        return False
    
    if interval and contract.is_ohlcv_subscribed:
        logger.info(f"ℹ️ {contract.short_hand} is already ohlcv subscribed")
        return True
    else:
        if contract.is_ohlcv_subscribed:
            unsubscribe_feed(contract, contract.ohlcv_interval)
        if (get_exchange_quotes and contract.is_ltp_subscribed) and (get_market_depth and contract.is_depth_subscribed):
            logger.info(f"ℹ️ {contract.short_hand} is already subscribed")
            return True

    try:
        # Convert ISO date to Breeze format using utility function
        expiry_date = convert_iso_to_breeze_date(contract.expiry_date)
        
        # Build parameters dictionary
        params = {
            "exchange_code" : contract.exchange_code,
            "stock_code" : contract.stock_code,
            "product_type" : "Options" if contract.product_type == "options" else "Futures",
            "expiry_date" : expiry_date,
            "strike_price" : str(contract.strike_price),
            "right" : "Call" if contract.right == "call" else "Put" if contract.right == "put" else "",
            "get_exchange_quotes" : get_exchange_quotes,
            "get_market_depth" : get_market_depth
        }

        # Conditionally add interval
        if interval is not None:
            params["interval"] = interval

        # Call API with unpacked parameters
        breeze.subscribe_feeds(**params)
        
        # Register this contract so the callback can find it
        contract_registry[contract.short_hand] = contract
            
        if interval:
            contract.is_ohlcv_subscribed = True
        else:
            if get_market_depth:
                contract.is_depth_subscribed = True
            if get_exchange_quotes:
                contract.is_ltp_subscribed = True

        logger.info(f"✅ Subscribed to feeds for {contract.short_hand}")
        return True
        
    except Exception as e:
        logger.error(f"❌ Failed to subscribe feeds for {contract.short_hand}: {e}")
        raise

def unsubscribe_feed(contract, interval=None, get_exchange_quotes=True, get_market_depth=False):
    #Unsubscribe from real-time feed for a contract
    
    if interval and not contract.is_ohlcv_subscribed:
        logger.info(f"ℹ️ {contract.short_hand} is not OHLCV subscribed")
        return
    else:
        if (get_exchange_quotes and not contract.is_ltp_subscribed) and (get_market_depth and not contract.is_depth_subscribed):
            logger.info(f"ℹ️ {contract.short_hand} is not subscribed")
            return

    try:
        # Convert ISO date to Breeze format using utility function
        expiry_date = convert_iso_to_breeze_date(contract.expiry_date)
        
        # Build parameters dictionary
        params = {
            "exchange_code" : contract.exchange_code,
            "stock_code" : contract.stock_code,
            "product_type" : "Options" if contract.product_type == "options" else "Futures",
            "expiry_date" : expiry_date,
            "strike_price" : str(contract.strike_price),
            "right" : "Call" if contract.right == "call" else "Put" if contract.right == "put" else "",
            "get_exchange_quotes" : get_exchange_quotes,
            "get_market_depth" : get_market_depth
        }

        # Conditionally add interval
        if interval is not None:
            params["interval"] = interval

        # Call API with unpacked parameters
        breeze.unsubscribe_feeds(**params)
        sleep(0.2)
        
        if interval:
            contract.is_ohlcv_subscribed = False 
            contract.ohlcv_data = []
        else: 
            if get_exchange_quotes:
                contract.is_ltp_subscribed = False
                # Reset the LTP and last update time to None when unsubscribing from price feed
                contract.ltp = None
                contract.last_update_time = None
            if get_market_depth:
                contract.is_depth_subscribed = False
                contract.best_bid_price = None
                contract.best_bid_qty = None
                contract.best_ask_price = None
                contract.best_ask_qty = None
                contract.depth_update_time = None

        # Unregister this contract
        if contract.short_hand in contract_registry and not contract.is_ltp_subscribed and not contract.is_ohlcv_subscribed and not contract.is_depth_subscribed:
            del contract_registry[contract.short_hand]
        logger.info(f"✅ Unsubscribed from feeds for {contract.short_hand}")
        
    except Exception as e:
        logger.error(f"❌ Failed to unsubscribe feeds for {contract.short_hand}: {e}")
        raise

def subscribe_multiple_feeds(contracts, interval=None, get_exchange_quotes=True, get_market_depth=False):
    #Subscribe to multiple feeds at once
    for contract in contracts:
        if not subscribe_feed(contract, interval, get_exchange_quotes, get_market_depth):
            return False
    return True

def unsubscribe_multiple_feeds(contracts, interval=None, get_exchange_quotes=True, get_market_depth=False):
    #Unsubscribe from multiple feeds at once
    for contract in contracts:
        unsubscribe_feed(contract, interval, get_exchange_quotes, get_market_depth)
    return
        
def unsubscribe_all_feed(interval=None, get_exchange_quotes=True, get_market_depth=False):
    # Unsubscribe from all web socket feeds
    unsubscribe_multiple_feeds(list(contract_registry.values()), interval, get_exchange_quotes, get_market_depth)
    
def create_date(date, month, year=2025):
    # Create ISO8601 date string for the given date and month
    if not (1 <= date <= 31) or not (1 <= month <= 12):
        raise ValueError("Invalid date or month")
    dt = datetime(year, month, date, 6, 0, 0)
    return dt.strftime("%Y-%m-%dT%H:%M:%S.000Z")
        
def get_current_expiry(target_weekday):
    # Get the date of the current or next occurrence of the specified weekday
    # 0-Monday, 1-Tuesday, 2-Wednesday,..., 6-Sunday
    today = datetime.now(pytz.timezone('Asia/Kolkata'))
    days_ahead = (target_weekday - today.weekday()) % 7  
    weekday = today + timedelta(days = days_ahead)
    print(weekday)
    return weekday.strftime("%Y-%m-%dT06:00:00.000Z")

class contract:
    # Class to represent a trading contract (futures or options)
    def __init__(self, stock_code, exchange_code, product_type, expiry_date, right="others", strike_price="0"):
        self.stock_code = stock_code
        self.exchange_code = exchange_code
        self.product_type = product_type
        self.expiry_date = expiry_date
        self.right = right
        self.strike_price = strike_price
        
        # Create short_hand with the Breeze format for expiry date
        if product_type == "options":
            self.short_hand = f"{stock_code}-{convert_iso_to_breeze_date(expiry_date)}-{strike_price}-{right}"
        else:
            self.short_hand = f"{stock_code}-FUT-{convert_iso_to_breeze_date(expiry_date)}"
        
        # Essential fields
        self.ltp = None
        self.last_update_time = None
        self.is_ltp_subscribed = False
        self.is_ohlcv_subscribed = False
        self.ohlcv_interval = None
        self.is_depth_subscribed = False
        
        # Market depth data
        self.best_bid_price = None
        self.best_bid_qty = None
        self.best_ask_price = None
        self.best_ask_qty = None
        self.depth_update_time = None
        
        # OHLCV candle storage
        self.ohlcv_data = []
        self.max_candles = 1000  # Maximum candles to store
        
        # Limit the size of the candle list
        if len(self.ohlcv_data) > self.max_candles:
            self.ohlcv_data = self.ohlcv_data[-self.max_candles:]
            
        # Store signals generated by ML function
        self.signal = None
        self.confidence = None
        
def generate_contracts(stock_code, expiry_date, start_strike, end_strike, interval, keyword=None, exchange_code="NFO", product_type="options"):
    contracts = []
    
    # Generate call and put option contracts for a range of strike prices
    for strike in range(start_strike, end_strike+1, interval):
        # Generating calls
        call_contract = contract(stock_code, exchange_code, product_type, expiry_date, "call", strike)
        globals()[f"{stock_code}{strike}CE{keyword if keyword else ''}"] = call_contract
        contracts.append(call_contract)
        
        # Generating puts
        put_contract = contract(stock_code, exchange_code, product_type, expiry_date, "put", strike)
        globals()[f"{stock_code}{strike}PE{keyword if keyword else ''}"] = put_contract
        contracts.append(put_contract)
            
        print(f"ℹ️ {stock_code}{strike}CE{keyword if keyword else ''}, {stock_code}{strike}PE{keyword if keyword else ''} generated.")
        
    return contracts

def clear():
    # Clear the output in Jupyter notebooks
    clear_output(wait=True)    

def place_fno_order(contract, action, quantity, count=1, price="0", stoploss="0", timeout=10):
    # Place F&O order(s) using ThreadPoolExecutor - simple and clean"""
    if action == "buy" and count*quantity > 4000:
        print("Maximum order limit exceeded")
        return
    
    order_params = {
        "stock_code": contract.stock_code,
        "exchange_code": contract.exchange_code,
        "product": contract.product_type,
        "action": action,
        "order_type": "market" if price == "0" else "limit",
        "stoploss": stoploss,
        "quantity": quantity,
        "price": price,
        "validity": "day",
        "validity_date": get_iso_date(),
        "disclosed_quantity": "0",
        "expiry_date": contract.expiry_date,
        "right": contract.right,
        "strike_price": contract.strike_price
    }
    
    def single_order():
        return breeze.place_order(**order_params)
    
    # Handle single order - use persistent pool
    if count == 1:
        try:
            future = ORDER_THREAD_POOL.submit(single_order)
            response = future.result(timeout=timeout)
            
            # Update positions if order was successful
            if response.get("Status") == 200:
                logger.info(f"✅ {contract.short_hand}-{action} order successful.")
            else:
                logger.error(f"❌ {contract.short_hand}-{action} order failed: {response.get('Error', 'Unknown error')}")
            
            return response
            
        except concurrent.futures.TimeoutError:
            logger.error(f"⏱️ {contract.short_hand}-{action} order timed out after {timeout}s")
            return {"Status": "Timeout", "Error": f"Order timed out after {timeout} seconds"}
    
    # Handle multiple orders - use persistent pool
    else:
        responses = []
        successful_orders = 0
        
        # Submit all orders to persistent pool
        futures = [ORDER_THREAD_POOL.submit(single_order) for _ in range(count)]
        
        # Collect results
        for i, future in enumerate(futures, 1):
            try:
                response = future.result(timeout=timeout)
                responses.append(response)
                
                # Update positions for successful orders
                if response.get("Status") == 200:
                    successful_orders += 1
                    logger.info(f"✅ {contract.short_hand}-{action} order {i}/{count} successful.")
                else:
                    logger.error(f"❌ {contract.short_hand}-{action} order {i}/{count} failed: {response.get('Error', 'Unknown error')}")
                    
            except concurrent.futures.TimeoutError:
                logger.error(f"⏱️ {contract.short_hand}-{action} order {i}/{count} timed out after {timeout}s")
                responses.append({"Status": "Timeout", "Error": f"Order {i} timed out after {timeout} seconds"})
        
        # Summary
        logger.info(f"📊 Order summary: {successful_orders}/{count} successful")
        
        return responses
    
def get_price(contract):
    # Get the current price for a contract
    try:
        response = breeze.get_quotes(
            stock_code = contract.stock_code,
            exchange_code = contract.exchange_code,
            expiry_date = contract.expiry_date,
            product_type = contract.product_type,
            right = contract.right,
            strike_price =contract.strike_price
        )
        if response.get("Success"):
            return float(response["Success"][0]["ltp"])
        else:
            logger.error(f"❌ API error in fetching {contract.short_hand}")
            print(response)
            return None
    except Exception as e:
        logger.error(f"❌ Error fetching {contract.short_hand} price: {e}")
        return None

def get_option_chain(stock_code, expiry_date, exchange_code="NFO"):
    expiry_date_display = convert_iso_to_breeze_date(expiry_date)
    logger.info(f"🔍 Fetching {stock_code} chain for {expiry_date_display}")
    
    option_data = {}
    for right in ["call", "put"]:
        chain = breeze.get_option_chain_quotes(
            stock_code=stock_code,
            exchange_code=exchange_code,
            product_type="options",
            expiry_date=expiry_date,
            right=right
        )
        if chain.get("Success"):
            option_data[right] = chain["Success"]
    
    if not option_data:
        logger.error("❌ Failed to fetch options")
        return None, None, None
    
    spot_price = float(option_data["call"][0]["spot_price"])
    
    # Process strikes
    strikes_data = {}
    for right, options in option_data.items():
        for opt in options:
            # Convert strike to integer by removing decimal part
            strike = int(float(opt["strike_price"]))
            if strike not in strikes_data:
                strikes_data[strike] = {"call": None, "put": None}
            strikes_data[strike][right] = opt
    
    all_strikes = sorted(strikes_data.keys())
    atm_strike = min(all_strikes, key=lambda s: abs(s - int(spot_price)))
    
    # Get display range
    atm_idx = all_strikes.index(atm_strike)
    num_strikes = min(20, len(all_strikes))
    start_idx = max(0, atm_idx - num_strikes // 2)
    end_idx = min(len(all_strikes), start_idx + num_strikes)
    if end_idx == len(all_strikes):
        start_idx = max(0, end_idx - num_strikes)
    display_strikes = all_strikes[start_idx:end_idx]
    
    # Display header with properly aligned columns
    print(f"\n{stock_code} Option Chain ({expiry_date_display}) - Spot: {int(spot_price)}")
    print("="*135)
    print(f"{'CALLS':<61} | {'STRIKE':<10} | {'PUTS':<65}")
    print(f"{'OI':<10}{'Ask':<12}{'Bid':<12}{'Chg%':<12}{'LTP':<15} | {'PRICE':<10} | {'LTP':<15}{'Chg%':<12}{'Bid':<12}{'Ask':<12}{'OI':<10}")
    print("-"*135)
    
    # Display rows with new column order
    for strike in display_strikes:
        row_data = {"call": {}, "put": {}}
        
        for right in ["call", "put"]:
            if strikes_data[strike][right]:
                opt = strikes_data[strike][right]
                row_data[right] = {
                    "oi": int(float(opt["open_interest"])) if opt["open_interest"] else "-",
                    "bid": float(opt["best_bid_price"]) if float(opt["best_bid_price"]) > 0 else "-",
                    "ask": float(opt["best_offer_price"]) if float(opt["best_offer_price"]) > 0 else "-",
                    "ltp": float(opt["ltp"]) if float(opt["ltp"]) > 0 else "-",
                    "chg": opt["ltp_percent_change"] if float(opt["ltp_percent_change"]) != 0 else "-"
                }
            else:
                row_data[right] = {"oi": "-", "bid": "-", "ask": "-", "ltp": "-", "chg": "-"}
        
        c = row_data["call"]
        p = row_data["put"]
        # Format strike price as integer (no decimals)
        strike_display = f"{strike}"
        print(f"{c['oi']:<10}{c['ask']:<12}{c['bid']:<12}{c['chg']:<12}{c['ltp']:<15} | {strike_display:<10} | {p['ltp']:<15}{p['chg']:<12}{p['bid']:<12}{p['ask']:<12}{p['oi']:<10}")
    
    print("="*135)
    
    # Return all strikes, spot price, atm strike, and raw option data
    return spot_price, atm_strike, option_data

def get_closest_price_pair(stock_code, expiry_date, exchange_code="NFO"):
    # Find two option contracts (one call, one put) with similar prices prioritizing contracts close to the ATM strike.
    
    # Get the option chain data
    spot_price, atm_strike, option_data = get_option_chain(
        stock_code, expiry_date, exchange_code
    )
    
    if not option_data or not atm_strike or not spot_price:
        logger.error("❌ Failed to fetch option chain")
        return None
    
    # Process options for pair finding
    call_options = []
    put_options = []
    
    # Process calls
    for opt in option_data["call"]:
        # Convert strike to integer
        strike = int(float(opt["strike_price"]))
        price = float(opt["ltp"]) if float(opt["ltp"]) > 0 else None
        
        # Use bid/ask midpoint if LTP is not available
        if price is None or price == 0:
            bid = float(opt["best_bid_price"]) if float(opt["best_bid_price"]) > 0 else 0
            ask = float(opt["best_offer_price"]) if float(opt["best_offer_price"]) > 0 else 0
            if bid > 0 and ask > 0:
                price = (bid + ask) / 2
            elif bid > 0:
                price = bid
            elif ask > 0:
                price = ask
        
        if price is not None and price > 0:
            call_options.append({
                "strike": strike,
                "price": price,
                "itm": strike <= int(spot_price),  # Call is ITM if strike <= spot
                "distance_from_atm": abs(strike - atm_strike)
            })
    
    # Process puts
    for opt in option_data["put"]:
        # Convert strike to integer
        strike = int(float(opt["strike_price"]))
        price = float(opt["ltp"]) if float(opt["ltp"]) > 0 else None
        
        # Use bid/ask midpoint if LTP is not available
        if price is None or price == 0:
            bid = float(opt["best_bid_price"]) if float(opt["best_bid_price"]) > 0 else 0
            ask = float(opt["best_offer_price"]) if float(opt["best_offer_price"]) > 0 else 0
            if bid > 0 and ask > 0:
                price = (bid + ask) / 2
            elif bid > 0:
                price = bid
            elif ask > 0:
                price = ask
        
        if price is not None and price > 0:
            put_options.append({
                "strike": strike,
                "price": price,
                "itm": strike >= int(spot_price),  # Put is ITM if strike >= spot
                "distance_from_atm": abs(strike - atm_strike)
            })
    
    # Sort by distance from ATM strike
    call_options.sort(key=lambda x: x["distance_from_atm"])
    put_options.sort(key=lambda x: x["distance_from_atm"])
    
    # Keep only the 5 closest strikes to ATM for both calls and puts
    near_atm_calls = call_options[:5] if len(call_options) > 5 else call_options
    near_atm_puts = put_options[:5] if len(put_options) > 5 else put_options
    
    # Find pairs with the most similar prices where at least one is ITM
    # and both are close to ATM
    best_pairs = []
    for call in near_atm_calls:
        for put in near_atm_puts:
            # Ensure at least one is ITM
            if call["itm"] or put["itm"]:
                price_diff = abs(call["price"] - put["price"])
                combined_distance = call["distance_from_atm"] + put["distance_from_atm"]
                
                # Calculate a score = price difference + normalized distance from ATM
                # Lower score is better
                score = price_diff + (combined_distance / spot_price) * 50
                
                best_pairs.append({
                    "call": call,
                    "put": put,
                    "price_diff": price_diff,
                    "combined_distance": combined_distance,
                    "score": score
                })
    
    if not best_pairs:
        logger.error("❌ No suitable option pairs found")
        return None
    
    # Sort by score (lower is better)
    best_pairs.sort(key=lambda x: x["score"])
    best_pair = best_pairs[0]
    
    call = best_pair["call"]
    put = best_pair["put"]
    
    logger.info(f"✅ Found option pair: {call['strike']}CE and {put['strike']}PE")
    logger.info(f"   Call: ₹{call['price']:.2f} ({('ITM' if call['itm'] else 'OTM')})")
    logger.info(f"   Put:  ₹{put['price']:.2f} ({('ITM' if put['itm'] else 'OTM')})")
    logger.info(f"   Price difference: ₹{best_pair['price_diff']:.2f}")
    
    # Create contract objects
    call_contract = contract(
        stock_code=stock_code,
        exchange_code=exchange_code,
        product_type="options",
        expiry_date=expiry_date,
        right="call",
        strike_price=call["strike"]
    )
    
    put_contract = contract(
        stock_code=stock_code,
        exchange_code=exchange_code,
        product_type="options",
        expiry_date=expiry_date,
        right="put",
        strike_price=put["strike"]
    )
    
    return [call_contract, put_contract]

def place_hedge_order(contract1, contract2, quantity, count = 1):
    # Place a hedge order (buy one contract, sell another) multiple times
    try:
        i = 0
        while i < count:
            response = place_fno_order(contract1, "buy", (quantity), price="0")
            if response.get("Status") == 200 :
                response = place_fno_order(contract2, "sell", (quantity), price="0")
                while response.get("Status") != 200 :
                    logger.info(f"🔄 Retrying...")
                    response = place_fno_order(contract2, "sell", (quantity), price="0")
            else:
                logger.info(f"🔄 Retrying...")
                i -= 1
            i += 1

    except KeyboardInterrupt:
        logger.info("Program terminated by user.")
    except Exception as e:
        logger.error(f"❌ Unexpected error in loop: {e}")

def hold_hedge(sell_contract, sell_quantity, sell_multiple, buy_contract, buy_quantity, buy_multiple, threshold=0, stop_loss=5):
    # Monitor and trade a contract based on price movements around a threshold
    try:
        # Place initial buy order if specified
        logger.info(f"✅ Placing initial buy order for {buy_contract.short_hand}")
        if buy_contract and buy_quantity and buy_multiple:
            place_fno_order(buy_contract, "buy", buy_quantity, buy_multiple)

        # Main trading loop
        logger.info(f"✅ Starting main trading loop for {sell_contract.short_hand}")
        hold_buy(sell_contract, sell_quantity, sell_multiple, threshold, stop_loss, position=1)
               
    except KeyboardInterrupt:
        logger.info("Operation terminated by user.")
    except Exception as e:
        logger.error(f"❌ Error in main loop: {e}")

def hold_buy(contract, quantity=1800, multiple=1, threshold=0, stop_loss=2, position=0):
    # Monitor and trade a contract based on price movements around a threshold
    try:
        count = 0
        running = True
        iterations = 0
        
        # Subscribe to feeds to get price updates
        if not subscribe_feed(contract):
            return
            
        # Wait for price data to be available
        while contract.ltp is None:
            sleep(0.1)
           
        # Get initial price if threshold not specified
        if threshold == 0:
            threshold = contract.ltp
            logger.info(f"✅ Using threshold: {threshold}")
            
        # Main trading loop
        while running:
            iterations += 1
           
            if iterations % 20 == 0:
                print(f"\rThreshold: {threshold} | Price: {contract.ltp} | Count: {count}    ", end="", flush=True)
           
            if contract.ltp is not None:
                if position == 0 and contract.ltp > threshold:
                    place_fno_order(contract, "buy", quantity, multiple)
                    position = 1
                    logger.info(f"✅ Bought at {contract.ltp}")
                elif position == 1 and contract.ltp < threshold-stop_loss:
                    place_fno_order(contract, "sell", quantity, multiple)
                    position = 0
                    count += 1
                    logger.info(f"✅ Sold at {contract.ltp}")
                if(contract.ltp > threshold+50):
                    threshold += 25
               
            sleep(0.02)
               
    except KeyboardInterrupt:
        logger.info("Operation terminated by user.")
    except Exception as e:
        logger.error(f"❌ Error in main loop: {e}")
    finally:
        running = False
        if position == 1:
            place_fno_order(contract, "sell", quantity, multiple)
        unsubscribe_feed(contract)  
         
def auto_switch_contracts(contracts, quantity=500, multiple=1, switch_diff=0, action="buy", max_loss=None):
    # Extract contracts
    if contracts and len(contracts) == 2:
        contract1 = contracts[0]
        contract2 = contracts[1]
    else:
        logger.error("❌ Could not start auto-switch strategy - invalid contracts")
        return
    
    # Calculate contract display names before the loop starts
    contract1_strike = contract1.strike_price
    contract1_right = "CE" if contract1.right == "call" else "PE"
    contract1_display = f"{contract1_strike}{contract1_right}"

    contract2_strike = contract2.strike_price
    contract2_right = "CE" if contract2.right == "call" else "PE"
    contract2_display = f"{contract2_strike}{contract2_right}"
    
    # Auto switch between two contracts based on price difference
    logger.info(f"🚀 Starting auto switch between {contract1.short_hand} and {contract2.short_hand}")

    # Make sure both contracts are subscribed for market depth
    if not (subscribe_feed(contract1, get_exchange_quotes=False, get_market_depth=True) and 
            subscribe_feed(contract2, get_exchange_quotes=False, get_market_depth=True)):
        logger.error("❌ Failed to subscribe to contract feeds")
        return
    
    # Wait for price data (only once, indefinitely)
    logger.info("⏳ Waiting for price data...")
    while contract1.best_bid_price is None or contract2.best_bid_price is None:
        sleep(0.05)
    
    # Determine which contract to use initially based on action
    if action == "buy":
        # When buying, choose the higher-priced contract
        current_contract = contract1 if contract1.best_bid_price >= contract2.best_bid_price else contract2
    else:  # action == "sell"
        # When selling, choose the lower-priced contract
        current_contract = contract2 if contract1.best_bid_price >= contract2.best_bid_price else contract1
    
    other_contract = contract2 if current_contract == contract1 else contract1
    
    # Place initial order
    logger.info(f"🔄 Initial {action}: {current_contract.short_hand} at price {current_contract.best_bid_price}")
    place_fno_order(current_contract, action, quantity, multiple)
    
    cover_action = "sell" if action == "buy" else "buy"
    switch_count = 0
    iterations = 0
    start_time = datetime.now()
    entry_price = current_contract.best_bid_price
    position_size = quantity * multiple
    total_pnl = 0
    max_pnl = 0
    
    # Monitor and switch loop
    try:
        while True:
            # Calculate P&L based on action
            if action == "buy":
                # For buy positions (profit when price goes up)
                current_pnl = (current_contract.best_bid_price - entry_price) * position_size
            else:
                # For sell positions (profit when price goes down)
                current_pnl = (entry_price - current_contract.best_bid_price) * position_size
            
            total_current_pnl = total_pnl + current_pnl
            
            # Check stop loss
            if max_loss and total_current_pnl < max_loss:
                logger.warning(f"⚠️ Max loss reached (₹{total_current_pnl:.2f})")
                place_fno_order(current_contract, cover_action, quantity, multiple)
                current_contract = None
                total_pnl += current_pnl
                break
            
            # Switching logic based on action
            should_switch = False
            if action == "buy":
                # When buying, switch when other contract becomes higher by threshold
                if other_contract.best_bid_price > current_contract.best_bid_price + switch_diff:
                    should_switch = True
            else:  # action == "sell"
                # When selling, switch when other contract becomes lower by threshold
                if current_contract.best_bid_price > other_contract.best_bid_price + switch_diff and current_contract.best_bid_price > entry_price:
                    should_switch = True
            
            if should_switch:
                # Cover current contract
                place_fno_order(current_contract, cover_action, quantity, multiple)
                
                # Update realized P&L
                if action == "buy":
                    exit_pnl = (current_contract.best_bid_price - entry_price) * position_size
                else:
                    exit_pnl = (entry_price - current_contract.best_bid_price) * position_size
                
                total_pnl += exit_pnl
                logger.debug(f"Realized P&L after switch: ₹{total_pnl:.2f}")
                
                # Enter other contract
                place_fno_order(other_contract, action, quantity, multiple)
                
                # Swap contracts and update entry price
                temp = current_contract
                current_contract = other_contract
                other_contract = temp
                entry_price = current_contract.best_bid_price
                
                logger.info(f"🔄 Switched to {current_contract.short_hand} at price {current_contract.best_bid_price}")
                switch_count += 1
            
            # Update max_pnl if current total P&L is greater
            if max_pnl < total_current_pnl:
                max_pnl = total_current_pnl
                logger.debug(f"New max P&L: ₹{max_pnl:.2f}")
            
            iterations +=1
            if iterations % 20 == 0:
                elapsed = (datetime.now() - start_time).total_seconds()
                
                print(f"\r{contract1_display}@{contract1.best_bid_price:.2f} | "
                  f"{contract2_display}@{contract2.best_bid_price:.2f} | "
                  f"P&L:{current_pnl:.0f}|{total_current_pnl:.0f}|{max_pnl:.0f} | "
                  f"SC:{switch_count} | {elapsed:.0f}s     ", 
                  end="", flush=True)
                
            sleep(0.02)
            
    except KeyboardInterrupt:
        logger.info("👋 Auto switch stopped by user")
    except Exception as e:
        logger.error(f"❌ Error in auto switch: {e}")
    finally:
        # Exit position and clean up
        try:
            if current_contract is not None:
                logger.info(f"🔄 Exiting position in {current_contract.short_hand}")
                place_fno_order(current_contract, cover_action, quantity, multiple)
                
                # Update final P&L
                if action == "buy":
                    final_pnl = (current_contract.best_bid_price - entry_price) * position_size
                else:
                    final_pnl = (entry_price - current_contract.best_bid_price) * position_size
                
                total_pnl += final_pnl
                # Check once more for max_pnl update
                total_current_pnl = total_pnl
                if max_pnl < total_current_pnl:
                    max_pnl = total_current_pnl
        except Exception as e:
            logger.error(f"❌ Error during exit: {e}")
            
        # Unsubscribe feeds
        try:
            unsubscribe_feed(contract1, get_exchange_quotes=False, get_market_depth=True)
            unsubscribe_feed(contract2, get_exchange_quotes=False, get_market_depth=True)
        except Exception as e:
            logger.error(f"❌ Error unsubscribing feeds: {e}")
        
        # Print final results
        duration = datetime.now() - start_time
        duration_str = f"{int(duration.total_seconds()//3600):02d}h:{int((duration.total_seconds()%3600)//60):02d}m:{int(duration.total_seconds()%60):02d}s"
        
        print("\n===== AUTO SWITCH RESULTS =====")
        print(f"Duration: {duration_str}")
        print(f"Switches: {switch_count}")
        print(f"Final P&L: ₹{total_pnl:.2f}")
        print(f"Maximum pnl reached: ₹{max_pnl:.2f}")
        
def test_websocket(contracts):
    # Test websocket connection by monitoring price updates for multiple contracts
    
    # Ensure contracts is a list even if a single contract was passed
    if not isinstance(contracts, list):
        contracts = [contracts]
    
    try:
        # Subscribe to all contracts
        for contract in contracts:
            if not subscribe_feed(contract):
                return
        
        print(f"\n✅ Started monitoring {len(contracts)} contracts")
        print("Press Ctrl+C to stop...\n")
        
        # Dictionary to store last update info for each contract
        last_updates = {contract.short_hand: {"price": None, "time": None} for contract in contracts}
        
        while True:
            clear()
            now = datetime.now()
            
            # Print header
            print(f"{'Contract':<25} {'Price':<10} {'Last Update':<15}")
            print("-" * 55)
            
            # Update and show data for each contract
            for contract in contracts:
                last_data = last_updates[contract.short_hand]
                
                # Check if price has changed
                if contract.ltp != last_data["price"]:
                    last_data["price"] = contract.ltp
                    last_data["time"] = now
                
                # Calculate time since last update
                time_since_update = ""
                if last_data["time"]:
                    seconds = (now - last_data["time"]).total_seconds()
                    time_since_update = f"{seconds:.1f}s ago"
                
                # Print contract info
                print(f"{contract.short_hand:<25} {contract.ltp if contract.ltp else 'No data':<10} {time_since_update:<15}")
            
            sleep(0.2)
            
    except KeyboardInterrupt:
        print("\nTest stopped by user")
    finally:
        # Make sure to unsubscribe from all feeds
        for contract in contracts:
            unsubscribe_feed(contract)

def place_exit_order(position_tuple):
    # Place a single exit order, splitting into chunks if necessary
    pos, action = position_tuple
    try:
        price_adjustment = 0
        price = float(pos["ltp"])
        qty = int(float(pos["quantity"]))
        if price > 15:
            price_adjustment = 15 if action == "buy" else -15
        price += price_adjustment
        max_qty = 600  # Maximum order size
        
        # Convert expiry date using utility function
        expiry_date = get_iso_date(pos["expiry_date"])
        
        while qty > 0:
            chunk_qty = min(qty, max_qty)
            qty -= chunk_qty

            response = breeze.square_off(
                exchange_code=str(pos["exchange_code"]),
                product=str(pos["product_type"]).lower(),
                stock_code=str(pos["stock_code"]),
                expiry_date=str(expiry_date),
                right=str(pos["right"]),
                strike_price=str(pos["strike_price"]),
                action=str(action),
                order_type="limit",  # Changed to limit since we're using price
                validity="day",
                stoploss="0",
                quantity=str(chunk_qty),
                price=str(price),  # Using adjusted price
                validity_date=str(get_iso_date()),
                trade_password="",
                disclosed_quantity="0"
            )
            
            if response.get("Status") == 200:
                logger.info(f"✅ {action.upper()}: {pos['stock_code']} {pos['strike_price']}{pos['right']} x {chunk_qty} @ {price}")
            else:
                logger.error(f"❌ Failed {action.upper()}: {pos['stock_code']} {pos['strike_price']}{pos['right']} x {chunk_qty} @ {price}: {response.get('Error', 'Unknown error')}")

    except Exception as e:
        logger.error(f"❌ Order failed - {pos['stock_code']}: {str(e)}")

def cancel_order(order):
    # Cancel a single order with logging
    try:
        breeze.cancel_order(order["exchange_code"], order["order_id"])
        logger.info(f"✅ Cancelled: {order['stock_code']} {order['strike_price']}{order['right']} (ID: {order['order_id']})")
    except Exception as e:
        logger.error(f"❌ Cancel failed - Order {order['order_id']}: {str(e)}")

def az5():
    # Emergency close all positions and cancel pending orders
    try:
        logger.info("🔄 AZ5 initiated")
        
        close_open_positions()
        
        cancel_pending_orders()
        
        logger.info("✅ AZ5 completed")
        
    except Exception as e:
        logger.error(f"❌ AZ5 failed: {e}")
        
def close_open_positions():
    # Close open positions
    positions_response = breeze.get_portfolio_positions()
    if positions_response.get("Error") is None:
        positions = positions_response.get("Success", [])
        if positions:
            active = [(p, "buy" if p["action"] == "Sell" else "sell") 
                     for p in positions 
                     if p["quantity"] != "0" and p["action"] != "NA"]
            
            if active:
                logger.info(f"🔄 Processing {len(active)} positions")
                with ThreadPoolExecutor(max_workers=30) as executor:
                    try:
                        # Submit all tasks
                        futures = [executor.submit(place_exit_order, pos) for pos in active]
                        
                        # Wait with timeout
                        for i, future in enumerate(futures, 1):
                            try:
                                future.result(timeout=10)  # 10 seconds per position
                                logger.info(f"✅ Position {i}/{len(active)} closed")
                            except concurrent.futures.TimeoutError:
                                logger.error(f"⏱️ Position {i}/{len(active)} timed out after 8s")
                            except Exception as e:
                                logger.error(f"❌ Error closing position {i}: {e}")
                                
                    except Exception as e:
                        logger.error(f"❌ Error in close_open_positions: {e}")
            else:
                logger.info("ℹ️ No active positions to close")
    else:
        logger.info("ℹ️ No positions available")

def cancel_pending_orders():
    # Cancel pending orders for both NFO and BFO
    for exchange_code in ["NFO", "BFO"]:
        orders_response = breeze.get_order_list(
            exchange_code=exchange_code,
            from_date=get_iso_date(),
            to_date=get_iso_date()
        )
        
        if orders_response.get("Error") is None:
            orders = orders_response.get("Success", [])
            pending = [o for o in orders if o["status"] in ["Ordered", "Requested"]]
            if pending:
                logger.info(f"🔄 Cancelling {len(pending)} {exchange_code} pending orders")
                with ThreadPoolExecutor(max_workers=30) as executor:
                    try:
                        # Submit all cancellation tasks
                        futures = [executor.submit(cancel_order, order) for order in pending]
                        
                        # Wait with timeout
                        for i, future in enumerate(futures, 1):
                            try:
                                future.result(timeout=10)  # 10 seconds per cancellation
                                logger.info(f"✅ Order {i}/{len(pending)} cancelled")
                            except concurrent.futures.TimeoutError:
                                logger.error(f"⏱️ Order {i}/{len(pending)} cancel timed out after 8s")
                            except Exception as e:
                                logger.error(f"❌ Error cancelling order {i}: {e}")
                                
                    except Exception as e:
                        logger.error(f"❌ Error in cancel_pending_orders: {e}")
            else:
                logger.info(f"ℹ️ No pending {exchange_code} orders to cancel")
        else:
            logger.info(f"ℹ️ No {exchange_code} orders available")

def get_historical_data(contract, start_datetime, end_datetime, interval="1minute"):
    # Get historical data for a contract for a specific time range
    
    # Parameters:
    # start_datetime & end_datetime in format "DD-MMM-YYYY HH:MM:SS" (e.g. "13-Mar-2025 09:30:00")
    # end_datetime - End time in format "DD-MMM-YYYY HH:MM:SS" (e.g. "13-Mar-2025 15:30:00")
    # interval - Data interval: "1second", "1minute", "5minute", "30minute" or "1day"
    
    # Returns:
    # pandas DataFrame with OHLCV data or None if there was an error
    try:
        # Convert input datetime strings to ISO format using existing utility functions
        start_time = get_iso_datetime(start_datetime)
        end_time = get_iso_datetime(end_datetime)
        
        logger.info(f"🔄 Fetching historical data for {contract.short_hand}")
        logger.info(f"🕒 Time range: {start_datetime} to {end_datetime}")
        logger.info(f"📊 Using interval: {interval}")
        
        # Get historical data
        response = breeze.get_historical_data_v2(
            interval=interval,
            from_date=start_time,
            to_date=end_time,
            stock_code=contract.stock_code,
            exchange_code=contract.exchange_code,
            product_type=contract.product_type,
            expiry_date=contract.expiry_date,
            right=contract.right,
            strike_price=contract.strike_price
        )
        
        # Check if successful
        if "Success" in response and len(response["Success"]) > 0:
            # Convert to DataFrame
            data = pd.DataFrame(response["Success"])
            
            # Format datetime and numeric columns
            data['datetime'] = pd.to_datetime(data['datetime'])
            for col in ['open', 'high', 'low', 'close']:
                data[col] = pd.to_numeric(data[col])
            
            # Convert volume and open interest to numeric and integers
            data['volume'] = pd.to_numeric(data['volume']).astype('int')
            if 'oi' in data.columns:
                data['oi'] = pd.to_numeric(data['oi']).astype('int')
            
            logger.info(f"✅ Retrieved {len(data)} data points")
            
            return data
        else:
            logger.error("❌ No data returned or error in API response")
            if "Error" in response:
                logger.error(f"❌ API Error: {response['Error']}")
            return None
            
    except Exception as e:
        logger.error(f"❌ Error fetching historical data: {e}")
        return None

def plot_historical_data(data, contract, show_volume=True):
    # Plot historical data for a contract
    try:
        import matplotlib.pyplot as plt
        from matplotlib.patches import Rectangle
        
        if data is None or len(data) == 0:
            logger.error("❌ No data to plot")
            return
            
        # Create figure with appropriate subplots
        if show_volume:
            fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(12, 8), 
                                         gridspec_kw={'height_ratios': [3, 1]})
        else:
            fig, ax1 = plt.subplots(1, 1, figsize=(12, 6))
        
        # Determine the interval for appropriate visualization
        is_intraday = (data['datetime'].max() - data['datetime'].min()).total_seconds() <= 24*60*60
        
        # Plot price data
        if len(data) > 1 and is_intraday:
            # Candlestick chart for intraday data
            for i, row in data.iterrows():
                # Candlestick color
                color = 'green' if row['close'] >= row['open'] else 'red'
                
                # Draw body
                body_height = abs(row['close'] - row['open'])
                body_bottom = min(row['open'], row['close'])
                rect = Rectangle((i-0.4, body_bottom), 0.8, body_height, 
                                color=color, alpha=0.8, zorder=2)
                ax1.add_patch(rect)
                
                # Draw wicks
                ax1.plot([i, i], [row['low'], body_bottom], color='black', linewidth=1, zorder=1)
                ax1.plot([i, i], [body_bottom+body_height, row['high']], color='black', linewidth=1, zorder=1)
            
            # Set x-ticks to datetime
            ax1.set_xticks(range(0, len(data), max(1, len(data)//10)))
            ax1.set_xticklabels(data['datetime'].iloc[::max(1, len(data)//10)].dt.strftime('%H:%M'), rotation=45)
        else:
            # Line chart for daily data or single point
            ax1.plot(data.index, data['close'], label='Close Price', color='blue', linewidth=1.5)
            
        ax1.set_title(f"{contract.short_hand} - {data['datetime'].min().strftime('%Y-%m-%d')} to {data['datetime'].max().strftime('%Y-%m-%d')}")
        ax1.set_ylabel('Price')
        ax1.grid(True, alpha=0.3)
        
        # Add volume subplot if requested
        if show_volume:
            ax2.bar(range(len(data)), data['volume'], color='navy', alpha=0.5)
            ax2.set_ylabel('Volume')
            ax2.grid(True, alpha=0.3)
            ax2.set_xticks(range(0, len(data), max(1, len(data)//10)))
            ax2.set_xticklabels(data['datetime'].iloc[::max(1, len(data)//10)].dt.strftime('%H:%M'), rotation=45)
        
        plt.tight_layout()
        plt.show()
        
    except Exception as e:
        logger.error(f"❌ Error plotting data: {e}")
        
def run_signal_generator(contracts, stop_event, min_candles=60, update_interval=1, max_workers=10):
    # Continuously update trading signals for contracts in parallel threads with optimized parameters.
    logger.info(f"Starting signal generator for {len(contracts)} contracts")
    
    # Function to process a single contract with optimized parameters
    def process_contract(contract):
        try:
            # Skip if insufficient data
            if not hasattr(contract, 'ohlcv_data') or not contract.ohlcv_data:
                return
            
            # Get data as DataFrame
            if isinstance(contract.ohlcv_data, list):
                if len(contract.ohlcv_data) < min_candles:
                    return
                df = pd.DataFrame(contract.ohlcv_data)
            else:
                if len(contract.ohlcv_data) < min_candles:
                    return
                df = contract.ohlcv_data.copy()
            
            # Verify columns
            if 'datetime' not in df.columns:
                return
            
            required_columns = ['open', 'high', 'low', 'close']
            if not all(col in df.columns for col in required_columns):
                return
            
            # Prepare data
            df['datetime'] = pd.to_datetime(df['datetime'])
            for col in required_columns:
                df[col] = pd.to_numeric(df[col])
            
            if 'volume' in df.columns:
                df['volume'] = pd.to_numeric(df['volume'])
            
            df = df.sort_values('datetime').reset_index(drop=True)
            df = df.tail(min_candles)
            
            # Calculate indicators
            indicators = {}
            indicators['current_price'] = df['close'].iloc[-1]
            indicators['price_change'] = df['close'].iloc[-1] / df['close'].iloc[-2] - 1
            
            for window in [5, 10, 20]:
                if len(df) >= window:
                    indicators[f'ma_{window}'] = df['close'].rolling(window=window).mean().iloc[-1]
                    indicators[f'close_to_ma_{window}'] = df['close'].iloc[-1] / indicators[f'ma_{window}'] - 1
            
            if len(df) >= 26:
                ema_12 = df['close'].ewm(span=12, adjust=False).mean()
                ema_26 = df['close'].ewm(span=26, adjust=False).mean()
                indicators['macd'] = ema_12.iloc[-1] - ema_26.iloc[-1]
                
                if len(df) >= 35:
                    macd_series = ema_12 - ema_26
                    indicators['macd_signal'] = macd_series.ewm(span=9, adjust=False).mean().iloc[-1]
            
            if len(df) >= 15:
                delta = df['close'].diff()
                gain = delta.where(delta > 0, 0)
                loss = -delta.where(delta < 0, 0)
                avg_gain = gain.rolling(window=14).mean().iloc[-1]
                avg_loss = loss.rolling(window=14).mean().iloc[-1]
                
                if avg_loss != 0:
                    rs = avg_gain / avg_loss
                    indicators['rsi'] = 100 - (100 / (1 + rs))
                else:
                    indicators['rsi'] = 100
            
            if len(df) >= 20:
                indicators['bb_middle'] = df['close'].rolling(window=20).mean().iloc[-1]
                indicators['bb_std'] = df['close'].rolling(window=20).std().iloc[-1]
                indicators['bb_upper'] = indicators['bb_middle'] + 2 * indicators['bb_std']
                indicators['bb_lower'] = indicators['bb_middle'] - 2 * indicators['bb_std']
                indicators['bb_position'] = (df['close'].iloc[-1] - indicators['bb_lower']) / (indicators['bb_upper'] - indicators['bb_lower'])
            
            # Calculate VWAP
            if 'volume' in df.columns:
                df['vwap'] = (df['volume'] * ((df['high'] + df['low'] + df['close']) / 3)).cumsum() / df['volume'].cumsum()
                indicators['vwap'] = df['vwap'].iloc[-1]
                indicators['price_to_vwap'] = df['close'].iloc[-1] / indicators['vwap'] - 1

            # Generate signal with weighted indicators
            # Weights based on observed performance: Increase BB and RSI, decrease MA
            weights = {
                'ma': 0.5,       # Reduced weight for Moving Averages
                'rsi': 1.5,      # Increased weight for RSI
                'macd': 1.0,     # Normal weight for MACD
                'bb': 1.5,       # Increased weight for Bollinger Bands
                'price': 1.0,    # Normal weight for Price Change
                'vwap': 1.0      # Normal weight for VWAP
            }
            
            buy_votes = sell_votes = 0
            total_weight = 0
            
            # Moving Average vote (reduced weight)
            if 'ma_5' in indicators and 'ma_10' in indicators:
                weight = weights['ma']
                total_weight += weight
                if indicators['close_to_ma_5'] > 0 and indicators['ma_5'] > indicators['ma_10']:
                    buy_votes += weight
                elif indicators['close_to_ma_5'] < 0 and indicators['ma_5'] < indicators['ma_10']:
                    sell_votes += weight
            
            # RSI vote (increased weight)
            if 'rsi' in indicators:
                weight = weights['rsi']
                total_weight += weight
                if indicators['rsi'] < 30:
                    buy_votes += weight
                elif indicators['rsi'] > 70:
                    sell_votes += weight
            
            # MACD vote
            if 'macd' in indicators and 'macd_signal' in indicators:
                weight = weights['macd']
                total_weight += weight
                if indicators['macd'] > indicators['macd_signal']:
                    buy_votes += weight
                elif indicators['macd'] < indicators['macd_signal']:
                    sell_votes += weight
            
            # Bollinger Bands vote (increased weight)
            if 'bb_position' in indicators:
                weight = weights['bb']
                total_weight += weight
                if indicators['bb_position'] < 0.2:
                    buy_votes += weight
                elif indicators['bb_position'] > 0.8:
                    sell_votes += weight
            
            # Price change vote
            if 'price_change' in indicators:
                weight = weights['price']
                total_weight += weight
                if indicators['price_change'] > 0.003:  # Reduced from 0.005 to 0.003
                    buy_votes += weight
                elif indicators['price_change'] < -0.003:  # Reduced from -0.005 to -0.003
                    sell_votes += weight
                    
            # VWAP vote
            if 'vwap' in indicators:
                weight = weights['vwap']
                total_weight += weight
                if df['close'].iloc[-1] > indicators['vwap']:
                    buy_votes += weight
                elif df['close'].iloc[-1] < indicators['vwap']:
                    sell_votes += weight
            
            # Calculate final signal
            buy_percentage = buy_votes / total_weight if total_weight > 0 else 0
            sell_percentage = sell_votes / total_weight if total_weight > 0 else 0
            
            # Lowered threshold criteria (from >0.5 to >0.45)
            if buy_percentage > 0.45 and buy_percentage > sell_percentage:
                signal = "BUY"
                confidence = buy_percentage
            elif sell_percentage > 0.45 and sell_percentage > buy_percentage:
                signal = "SELL"
                confidence = sell_percentage
            else:
                signal = "HOLD"
                confidence = 1 - (buy_percentage + sell_percentage)
            
            # Update contract
            old_signal = getattr(contract, 'signal', None)
            contract.signal = signal
            contract.confidence = confidence
            
            # Log significant changes
            if old_signal != signal and confidence > 0.5:
                contract_id = getattr(contract, 'short_hand', f"{contract.stock_code}-{contract.strike_price}-{contract.right}")
                logger.info(f"🔔 {contract_id}: Signal changed {old_signal} → {signal} with {confidence:.2f} confidence @ ₹{contract.ltp}")
                
        except Exception as e:
            contract_id = getattr(contract, 'short_hand', str(contract))
            logger.error(f"Error analyzing {contract_id}: {str(e)}")
    
    # Main loop
    while not stop_event.is_set():
        try:
            # Create a thread pool
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                # Submit all contracts for processing
                executor.map(process_contract, contracts)
            
            # Sleep before next update
            sleep(update_interval)
            
        except Exception as e:
            logger.error(f"Error in signal generator: {str(e)}")
            sleep(update_interval)
    
    logger.info("Signal generator stopped")

def auto_trade_signals(contracts, quantity, confidence_threshold=0.5, candle_interval="1second", 
                      min_price=100, max_price=150, stop_loss=0.04, take_profit=0.05, stop_event=None):
    # Automatically trade based on signals generated for a list of contracts.
    
    # Create stop event if not provided
    if stop_event is None:
        stop_event = threading.Event()

    # Signal generator stop event
    signal_stop_event = threading.Event()

    # Flag to indicate if we're currently monitoring a position
    currently_monitoring = threading.Event()

    # Subscribe to feeds for all contracts
    logger.info(f"🔄 Subscribing to feeds for {len(contracts)} contracts")
    if not subscribe_multiple_feeds(contracts) or not subscribe_multiple_feeds(contracts, interval=candle_interval):
        return
    
    # Start signal generator in a separate thread
    signal_thread = threading.Thread(
        target=run_signal_generator,
        args=(contracts, signal_stop_event, 20, 1, 10)  # min_candles=20, update_interval=1s, max_workers=10
    )
    signal_thread.daemon = True
    signal_thread.start()
    logger.info("✅ Signal generator started")
    logger.info(f"✅ Trading contracts with price between ₹{min_price} and ₹{max_price}")
    logger.info(f"✅ Using stop-loss: {stop_loss*100}%, take-profit: {take_profit*100}%")

    # Function to monitor a single position until signal changes
    def monitor_position(contract):
        try:
            # Set flag that we're monitoring a position
            currently_monitoring.set()
            
            entry_price = contract.ltp
            entry_time = datetime.now()
            
            logger.info(f"🔍 Monitoring position for {contract.short_hand} entered at ₹{entry_price}")
            
            while not stop_event.is_set():
                # Exit conditions
                '''if contract.signal != "BUY" and contract.confidence >= confidence_threshold:
                    logger.info(f"📊 Signal changed to {contract.signal} with confidence {contract.confidence:.2f}")
                    logger.info(f"🔄 Exiting position for {contract.short_hand} at ₹{contract.ltp}")
                    
                    # Place exit order
                    place_fno_order(contract, "sell", quantity)
                    break'''
                
                # Safety net: Set a stop-loss or take-profi1
                if contract.ltp is not None and entry_price is not None:
                    # Calculate current P&L percentage
                    current_pnl_percent = ((contract.ltp / entry_price) - 1) * 100
                    
                    # Stop-loss: Exit if loss exceeds stop-loss percentage
                    if current_pnl_percent < -stop_loss * 100:
                        logger.info(f"⚠️ Stop-loss triggered for {contract.short_hand} at ₹{contract.ltp}")
                        place_fno_order(contract, "sell", quantity)
                        break
                    
                    # Take-profit: Exit if profit exceeds take-profit percentage
                    if current_pnl_percent > take_profit * 100:
                        logger.info(f"💹 Take-profit triggered for {contract.short_hand} at ₹{contract.ltp}")
                        place_fno_order(contract, "sell", quantity)
                        break
                
                # Exit if position held for more than 30 minutes
                time_held = datetime.now() - entry_time
                if time_held.total_seconds() > 1800:  # 30 minutes in seconds
                    logger.info(f"⏱️ Max holding time reached for {contract.short_hand}, exiting at ₹{contract.ltp}")
                    place_fno_order(contract, "sell", quantity)
                    break
                
                # If price moves out of our range during monitoring, exit the position
                if contract.ltp < min_price or contract.ltp > max_price:
                    logger.info(f"⚠️ Price moved outside trading range ({min_price}-{max_price}) for {contract.short_hand}, exiting at ₹{contract.ltp}")
                    place_fno_order(contract, "sell", quantity)
                    break
                
                # Sleep to avoid excessive CPU usage
                sleep(1)
                
        except Exception as e:
            logger.error(f"❌ Error monitoring position for {contract.short_hand}: {e}")
            
            # Attempt emergency exit if there was an error
            try:
                logger.info(f"🔄 Emergency exit for {contract.short_hand}")
                place_fno_order(contract, "sell", quantity)
            except:
                logger.error(f"❌ Failed emergency exit for {contract.short_hand}")
        finally:
            # Clear the monitoring flag so we can move to the next contract
            currently_monitoring.clear()

    # Main trading loop
    try:
        logger.info("🚀 Starting trading system - one contract at a time")
        
        while not stop_event.is_set():
            # Only look for a new contract if we're not currently monitoring one
            if not currently_monitoring.is_set():
                # Look for the contract with the highest buy confidence
                best_contract = None
                best_confidence = confidence_threshold  # Minimum threshold
                
                for contract in contracts:
                    # Skip if no LTP or signal available
                    if not hasattr(contract, 'ltp') or contract.ltp is None or not hasattr(contract, 'signal'):
                        continue
                    
                    # Skip if price is outside our trading range
                    if contract.ltp < min_price or contract.ltp > max_price:
                        continue
                    
                    # Check if this is a strong buy signal
                    if contract.signal == "BUY" and hasattr(contract, 'confidence'):
                        if contract.confidence > best_confidence:
                            best_contract = contract
                            best_confidence = contract.confidence
                
                # If we found a good candidate, trade it
                if best_contract is not None:
                    logger.info(f"🔔 Found BUY signal for {best_contract.short_hand} with confidence {best_confidence:.2f} at price ₹{best_contract.ltp}")
                    
                    # Place buy order
                    response = place_fno_order(best_contract, "buy", quantity)
                    
                    # Check if order was successful
                    if response and response.get("Status") == 200:
                        logger.info(f"✅ Buy order placed for {best_contract.short_hand} at ₹{best_contract.ltp}")
                        
                        # Start monitoring thread for this position
                        monitor_thread = threading.Thread(
                            target=monitor_position,
                            args=(best_contract,)
                        )
                        monitor_thread.daemon = True
                        monitor_thread.start()
                    else:
                        logger.error(f"❌ Failed to place buy order for {best_contract.short_hand}")
                else:
                    # Log status periodically
                    candidate_count = sum(1 for c in contracts 
                                      if hasattr(c, 'ltp') and c.ltp is not None 
                                      and min_price <= c.ltp <= max_price)
                    
                    if candidate_count > 0:
                        logger.info(f"ℹ️ No strong BUY signals found among {candidate_count} contracts in price range ₹{min_price}-₹{max_price}")
            
            # Sleep to avoid excessive polling
            sleep(2)
            
    except KeyboardInterrupt:
        logger.info("👋 Trading system stopped by user")
    except Exception as e:
        logger.error(f"❌ Error in trading system: {e}")
    finally:
        # Cleanup
        try:
            # Stop signal generator
            signal_stop_event.set()
            
            # If we're currently monitoring a position, try to close it
            if currently_monitoring.is_set():
                for contract in contracts:
                    if hasattr(contract, 'ltp') and contract.ltp is not None:
                        logger.info(f"🔄 Attempting to close any open position for {contract.short_hand}")
                        place_fno_order(contract, "sell", quantity)
            
            # Unsubscribe from all feeds
            logger.info(f"🔄 Unsubscribing from feeds for {len(contracts)} contracts")
            unsubscribe_multiple_feeds(contracts)
            unsubscribe_multiple_feeds(contracts, interval=candle_interval)
            
            logger.info("✅ Trading system shutdown complete")
        except Exception as e:
            logger.error(f"❌ Error during cleanup: {e}")
   
def get_pnl(start_date = get_iso_date(), end_date = get_iso_date()):
    # Get trade book and calculate P&L
    all_trades = []
    
    # Get trades for NFO and BFO
    for exchange_code in ["NFO", "BFO"]:
        response = breeze.get_trade_list(
            exchange_code=exchange_code,
            from_date=start_date,
            to_date=end_date
        )
        
        if response.get("Status") == 200 and response.get("Success"):
            all_trades.extend(response.get("Success", []))
            print(f"Retrieved {int(len(response.get('Success', []))/2)} trades for {exchange_code}")
    
    if not all_trades:
        print("No trades found for the specified date range")
        return
    
    # Convert to DataFrame
    trades_df = pd.DataFrame(all_trades)
    
    # Fix data types and handle empty values
    for col in ['quantity', 'average_cost']:
        trades_df[col] = pd.to_numeric(trades_df[col], errors='coerce')
    
    # Convert brokerage_amount and total_taxes to numeric, replacing empty or zero values with NaN
    trades_df['brokerage_amount'] = pd.to_numeric(trades_df['brokerage_amount'], errors='coerce')
    trades_df['total_taxes'] = pd.to_numeric(trades_df['total_taxes'], errors='coerce')
    
    # Check if we have valid brokerage and tax data
    has_valid_brokerage = (trades_df['brokerage_amount'] > 0).any()
    has_valid_taxes = (trades_df['total_taxes'] > 0).any()
    
    # Calculate trade value
    trades_df['trade_value'] = trades_df['quantity'] * trades_df['average_cost']
    
    # Calculate brokerage if not provided
    if not has_valid_brokerage:
        def calculate_brokerage(row):
            if row['product_type'] == 'Options':
                # Flat Rs. 20 per executed order for options
                return 20
            else:  # Futures
                # 0.03% or Rs. 20/executed order whichever is lower
                percentage_brokerage = row['trade_value'] * 0.0003
                return min(percentage_brokerage, 20)
        
        trades_df['brokerage_amount'] = trades_df.apply(calculate_brokerage, axis=1)
    
    # Calculate STT (Securities Transaction Tax)
    def calculate_stt(row):
        if row['action'] != 'Sell':
            return 0  # STT only on sell side for F&O
        
        if row['product_type'] == 'Options':
            # 0.1% on sell side (premium) for options
            return row['trade_value'] * 0.001
        else:  # Futures
            # 0.02% on the sell side for futures
            return row['trade_value'] * 0.0002
    
    trades_df['stt'] = trades_df.apply(calculate_stt, axis=1)
    
    # Calculate transaction charges
    def calculate_transaction_charges(row):
        if row['exchange_code'] == 'NFO' or row['exchange_code'] == 'NSE':
            if row['product_type'] == 'Options':
                # NSE: 0.03503% (on premium) for options
                return row['trade_value'] * 0.0003503
            else:  # Futures
                # NSE: 0.00173% for futures
                return row['trade_value'] * 0.0000173
        else:  # BSE
            if row['product_type'] == 'Options':
                # BSE: 0.0325% (on premium) for options
                return row['trade_value'] * 0.000325
            else:  # Futures
                # BSE: 0 for futures
                return 0
    
    trades_df['transaction_charges'] = trades_df.apply(calculate_transaction_charges, axis=1)
    
    # Calculate SEBI charges - ₹10 per crore (0.0000001%)
    trades_df['sebi_charges'] = trades_df['trade_value'] * 0.0000001
    
    # Calculate stamp duty
    def calculate_stamp_duty(row):
        if row['action'] != 'Buy':
            return 0  # Stamp duty only on buy side
        
        if row['product_type'] == 'Options':
            # 0.003% or ₹300 / crore on buy side for options
            return row['trade_value'] * 0.00003
        else:  # Futures
            # 0.002% or ₹200 / crore on buy side for futures
            return row['trade_value'] * 0.00002
    
    trades_df['stamp_duty'] = trades_df.apply(calculate_stamp_duty, axis=1)
    
    # Calculate GST - 18% on (brokerage + SEBI charges + transaction charges)
    trades_df['gst'] = (trades_df['brokerage_amount'] + trades_df['sebi_charges'] + trades_df['transaction_charges']) * 0.18
    
    # Calculate total taxes if not provided or invalid
    if not has_valid_taxes:
        trades_df['total_taxes'] = trades_df['stt'] + trades_df['transaction_charges'] + trades_df['sebi_charges'] + trades_df['stamp_duty'] + trades_df['gst']
    
    # Calculate raw P&L
    trades_df['raw_pnl'] = trades_df.apply(
        lambda row: row['trade_value'] if row['action'] == 'Sell' else -row['trade_value'], 
        axis=1
    )
    
    # Calculate final P&L
    trades_df['total_costs'] = trades_df['brokerage_amount'] + trades_df['total_taxes']
    trades_df['final_pnl'] = trades_df['raw_pnl'] - trades_df['total_costs']
    
    # Group by contract
    trades_df['contract_id'] = trades_df.apply(
        lambda row: f"{row['stock_code']}_{row['expiry_date']}_{row['strike_price']}_{row['right']}", 
        axis=1
    )
    
    contract_summary = trades_df.groupby(['exchange_code', 'contract_id']).agg({
        'stock_code': 'first',
        'expiry_date': 'first',
        'strike_price': 'first',
        'right': 'first',
        'final_pnl': 'sum',
        'trade_value': 'sum',
        'brokerage_amount': 'sum',
        'total_taxes': 'sum'
    }).reset_index()
    
    # Print contract-wise P&L by exchange
    print("\n===== CONTRACT-WISE P&L =====")
    
    # First NFO contracts
    print("\nNFO CONTRACTS:")
    nfo_contracts = contract_summary[contract_summary['exchange_code'] == 'NFO'].sort_values('final_pnl', ascending=False)
    for _, row in nfo_contracts.iterrows():
        print(f"{row['stock_code']} {row['expiry_date']} {row['strike_price']} {row['right']}: ₹ {row['final_pnl']:,.2f}")
    
    # Then BFO contracts
    print("\nBFO CONTRACTS:")
    bfo_contracts = contract_summary[contract_summary['exchange_code'] == 'BFO'].sort_values('final_pnl', ascending=False)
    for _, row in bfo_contracts.iterrows():
        print(f"{row['stock_code']} {row['expiry_date']} {row['strike_price']} {row['right']}: ₹ {row['final_pnl']:,.2f}")
    
    # Calculate overall totals
    gross_pnl = trades_df['raw_pnl'].sum()
    brokerage = trades_df['brokerage_amount'].sum()
    total_taxes = trades_df['total_taxes'].sum()
    net_pnl = gross_pnl - brokerage - total_taxes
    
    # Get tax component totals
    stt = trades_df['stt'].sum() if 'stt' in trades_df else total_taxes * 0.65
    transaction = trades_df['transaction_charges'].sum() if 'transaction_charges' in trades_df else total_taxes * 0.15
    stamp = trades_df['stamp_duty'].sum() if 'stamp_duty' in trades_df else total_taxes * 0.10
    gst = trades_df['gst'].sum() if 'gst' in trades_df else brokerage * 0.18
    
    # Print overall summary
    print("\n===== TRADE P&L SUMMARY =====")
    print(f"Total Trades: {int(len(trades_df)/2)}")
    print(f"Gross P&L (Before Charges): ₹ {gross_pnl:,.2f}")
    print("\nCHARGES BREAKDOWN:")
    print(f"Brokerage: ₹{brokerage:,.2f}")
    print(f"STT: ₹{stt:,.2f}")
    print(f"Transaction Charges: ₹{transaction:,.2f}")
    print(f"Stamp Duty: ₹{stamp:,.2f}")
    print(f"GST: ₹{gst:,.2f}")
    print(f"Total Charges: ₹{(brokerage + total_taxes):,.2f}")
    print("\nNET P&L (After Charges): ₹{:,.2f}".format(net_pnl))
   
def analyse_trades(start_date=None, end_date=None, filename=None):
    """Create Excel analysis of trades with buy/sell pair matching"""
    import pandas as pd
    from datetime import datetime
    
    # Set defaults
    if start_date is None: start_date = get_iso_date()
    if end_date is None: end_date = get_iso_date()
    if filename is None: filename = f"trade_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    # Get trades from both exchanges
    all_trades = []
    for exchange_code in ["NFO", "BFO"]:
        try:
            response = breeze.get_trade_list(exchange_code=exchange_code, from_date=start_date, to_date=end_date)
            if response.get("Status") == 200 and response.get("Success"):
                all_trades.extend(response.get("Success", []))
        except Exception as e:
            logger.error(f"❌ Error fetching {exchange_code} trades: {e}")
    
    if not all_trades:
        logger.error("❌ No trades found")
        return None
    
    # Convert to DataFrame and clean data
    df = pd.DataFrame(all_trades)
    df[['quantity', 'average_cost']] = df[['quantity', 'average_cost']].apply(pd.to_numeric, errors='coerce')
    df['contract_descriptor'] = df.apply(lambda r: f"OPT-{r['stock_code']}-{r['expiry_date']}-{r['strike_price']}-{'CE' if r['right'].lower() == 'call' else 'PE'}", axis=1)
    df = df.sort_values('order_id')
    
    # Combine consecutive trades with same action
    combined_trades = []
    for name, group in df.groupby(['stock_code', 'expiry_date', 'strike_price', 'right', 'exchange_code']):
        group = group.sort_values('order_id').reset_index(drop=True)
        i = 0
        while i < len(group):
            current = group.iloc[i]
            qty, value = current['quantity'], current['quantity'] * current['average_cost']
            j = i + 1
            while j < len(group) and group.iloc[j]['action'] == current['action']:
                next_trade = group.iloc[j]
                qty += next_trade['quantity']
                value += next_trade['quantity'] * next_trade['average_cost']
                j += 1
            
            combined = current.copy()
            combined['quantity'] = qty
            combined['average_cost'] = value / qty
            combined_trades.append(combined)
            i = j
    
    # Match buy/sell pairs using FIFO
    trades_df = pd.DataFrame(combined_trades)
    analysis = []
    
    for name, group in trades_df.groupby(['stock_code', 'expiry_date', 'strike_price', 'right', 'exchange_code']):
        buys = group[group['action'] == 'Buy'].sort_values('order_id')
        sells = group[group['action'] == 'Sell'].sort_values('order_id')
        
        buy_queue = []
        for _, buy in buys.iterrows():
            buy_queue.append({
                'order_ref': buy['order_id'],
                'trade_date': buy['trade_date'],
                'contract_descriptor': buy['contract_descriptor'],
                'exchange': buy['exchange_code'],
                'buy_price': round(float(buy['average_cost']), 2),
                'remaining_qty': int(buy['quantity'])
            })
        
        for _, sell in sells.iterrows():
            sell_qty = int(sell['quantity'])
            sell_price = round(float(sell['average_cost']), 2)
            
            while sell_qty > 0 and buy_queue:
                buy_entry = buy_queue[0]
                match_qty = min(sell_qty, buy_entry['remaining_qty'])
                diff = round(sell_price - buy_entry['buy_price'], 2)
                
                analysis.append({
                    'Order Ref.': buy_entry['order_ref'],
                    'Trade Date': buy_entry['trade_date'],
                    'Contract Descriptor': buy_entry['contract_descriptor'],
                    'Exch': buy_entry['exchange'],
                    'Qty': int(match_qty),
                    'Buy Price': buy_entry['buy_price'],
                    'Sell Price': sell_price,
                    'Diff.': diff,
                    'Value': round(diff * match_qty, 2)
                })
                
                sell_qty -= match_qty
                buy_entry['remaining_qty'] -= match_qty
                if buy_entry['remaining_qty'] == 0:
                    buy_queue.pop(0)
    
    # Create Excel file
    analysis_df = pd.DataFrame(analysis).sort_values('Value')
    
    try:
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            analysis_df.to_excel(writer, sheet_name='Trade Analysis', index=False)
            
            # Format sheet
            ws = writer.sheets['Trade Analysis']
            widths = {'A': 25, 'B': 12, 'C': 45, 'D': 8, 'E': 8, 'F': 12, 'G': 12, 'H': 10, 'I': 12}
            for col, width in widths.items():
                ws.column_dimensions[col].width = width
            
            # Color coding
            from openpyxl.styles import PatternFill
            red_fill = PatternFill(start_color='FFCCCB', end_color='FFCCCB', fill_type='solid')
            green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
            
            for row in range(2, len(analysis_df) + 2):
                cell = ws[f'I{row}']
                if cell.value < 0:
                    cell.fill = red_fill
                elif cell.value > 0:
                    cell.fill = green_fill
        
        total_trades = len(analysis_df)
        total_pnl = analysis_df['Value'].sum()
        profitable = len(analysis_df[analysis_df['Value'] > 0])
        
        logger.info(f"✅ Analysis exported: {filename}")
        logger.info(f"📊 Trades: {total_trades} | P&L: ₹{total_pnl:,.2f} | Win Rate: {(profitable/total_trades*100):.1f}%")
        return filename
        
    except Exception as e:
        logger.error(f"❌ Error creating Excel: {e}")
        return None
    
def auto_start(function, ist_start_hour=9, ist_start_min=15):
    # Timer that checks if the current India time matches the specified hour and minute and executes the provided function at that time.
    
    target_time = f"{ist_start_hour:02d}:{ist_start_min:02d}"
    logger.info(f"🕒 Market timer started - waiting for India time {target_time}")
    
    try:
        while True:
            # Get current time in UTC
            utc_now = datetime.now(timezone.utc)
            
            # Convert to India time (UTC+5:30)
            india_now = utc_now + timedelta(hours=5, minutes=30)
            india_hour = india_now.hour
            india_minute = india_now.minute
            
            # Get local time for display
            local_now = datetime.now()
            local_time_str = local_now.strftime("%H:%M:%S")
            
            # Display current status (update every 30 seconds)
            if local_now.second % 30 == 0:
                clear()
                print(f"{'='*50}")
                print(f"🕒 MARKET TIMER - Local time: {local_time_str}")
                print(f"🕒 India time: {india_hour:02d}:{india_minute:02d}")
                print(f"{'='*50}")
                print(f"Waiting for India time {target_time}")
                print(f"{'='*50}")
            
            # Check if it's the target time in India
            if india_hour == ist_start_hour and india_minute == ist_start_min:
                print(f"⏰ TARGET TIME {target_time} REACHED - Executing function")
                function()
                break
            
            # Sleep briefly to avoid high CPU usage
            sleep(1)
            
    except KeyboardInterrupt:
        print("\n👋 Market timer stopped by user")
    except Exception as e:
        logger.error(f"❌ Error in market timer: {e}")
        import traceback
        traceback.print_exc()
        
# Main execution
if __name__ == "__main__":
    # Initiate logger
    logger = log()
    logger.info(f"✅ Current IST time: {get_ist_time()}")
    
    # Get API connection
    get_breeze()

    # Give kill alias to az5
    az5 = kill = az5
    
    x=generate_contracts("NIFTY", create_date(10,7), 22000, 26000, 50)
    y=generate_contracts("BSESEN", create_date(8, 7), 70000, 85000, 100, exchange_code="BFO")
